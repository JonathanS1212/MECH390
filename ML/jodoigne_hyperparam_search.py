"""
Jodoigne Hyperparameter Grid Search
====================================
Sweeps over:
  - Loss function : nn.MSELoss | nn.L1Loss | nn.HuberLoss
  - Dropout       : 0.10 | 0.15
  - Max epochs    : 2000 | 3000 | 4000
  - Optimizer     : AdamW | Adam

For every combination the model is trained and the per-target R²
plus the mean R² are stored.  Results are written to
  hyperparameter_results.xlsx   (same folder as this script)
"""

# ─────────────────────────────────────────────────────────────────────────────
#  Imports
# ─────────────────────────────────────────────────────────────────────────────
import os
import itertools
import time

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import openpyxl as pxl
from torch.utils.data import Dataset, DataLoader
from sklearn.model_selection import train_test_split
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ─────────────────────────────────────────────────────────────────────────────
#  Fixed hyper-parameters  (unchanged from Jodoigne.ipynb)
# ─────────────────────────────────────────────────────────────────────────────
ALPHA       = 0.5          # mixup bell-curve coefficient
SEED        = 42
PATIENCE    = 100
BATCH_SIZE  = 32
FILE_NAME   = "FinalB"     # Excel data file (in ./Data/ subfolder)

# Grid to sweep
LOSS_FNS   = {
    "MSELoss"   : nn.MSELoss(),
    "L1Loss"    : nn.L1Loss(),
    "HuberLoss" : nn.HuberLoss(),
}
DROPOUTS   = [0.10, 0.15]
EPOCH_LIST = [2000, 3000, 4000]
OPTIMIZERS = ["AdamW", "Adam"]

# ─────────────────────────────────────────────────────────────────────────────
#  Dataset definitions  (identical to notebook)
# ─────────────────────────────────────────────────────────────────────────────
class Training(Dataset):
    def __init__(self, x_mix, y_mix):
        self.x_mix = x_mix
        self.y_mix = y_mix
    def __len__(self):        return self.x_mix.shape[0]
    def __getitem__(self, i): return self.x_mix[i], self.y_mix[i]

class Validation(Dataset):
    def __init__(self, x, y):
        self.x = x; self.y = y
    def __len__(self):        return self.x.shape[0]
    def __getitem__(self, i): return self.x[i], self.y[i]

class Testing(Dataset):
    def __init__(self, x, y):
        self.x = x; self.y = y
    def __len__(self):        return self.x.shape[0]
    def __getitem__(self, i): return self.x[i], self.y[i]


# ─────────────────────────────────────────────────────────────────────────────
#  Helper functions  (identical to notebook)
# ─────────────────────────────────────────────────────────────────────────────
def get_dataloaders(train_ds, val_ds, test_ds, batch_size, num_workers, seed):
    g = torch.Generator().manual_seed(seed)
    train_loader = DataLoader(
        train_ds, batch_size=batch_size, shuffle=True,
        drop_last=True, num_workers=num_workers, generator=g,
    )
    val_loader  = DataLoader(val_ds,  batch_size=batch_size, shuffle=False,
                             drop_last=False, num_workers=num_workers)
    test_loader = DataLoader(test_ds, batch_size=batch_size, shuffle=False,
                             drop_last=False, num_workers=num_workers)
    return train_loader, val_loader, test_loader


def split_data(x, y, seed, train_frac=0.70, val_frac=0.20):
    n          = len(x)
    test_frac  = 1.0 - train_frac - val_frac
    indices    = np.arange(n)
    train_idx, valtest_idx = train_test_split(
        indices, test_size=val_frac + test_frac, random_state=seed)
    relative_test = test_frac / (val_frac + test_frac)
    val_idx, test_idx = train_test_split(
        valtest_idx, test_size=relative_test, random_state=seed)
    return (torch.tensor(train_idx),
            torch.tensor(val_idx),
            torch.tensor(test_idx))


def mixup(x, y, alpha=ALPHA, seed=SEED):
    np.random.seed(seed)
    N     = x.size(0)
    lam   = np.random.beta(alpha, alpha, size=N)
    lam   = np.maximum(lam, 1 - lam)
    lam_t = torch.tensor(lam, dtype=torch.float32).unsqueeze(1)
    idx_p = torch.randperm(N)
    return lam_t * x + (1 - lam_t) * x[idx_p], \
           lam_t * y + (1 - lam_t) * y[idx_p]


def train_epoch(model, loader, optimizer, device, criterion):
    model.train()
    total = 0.0
    for xb, yb in loader:
        xb, yb = xb.to(device), yb.to(device)
        optimizer.zero_grad()
        loss = criterion(model(xb), yb)
        loss.backward()
        torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
        optimizer.step()
        total += loss.item() * len(xb)
    return total / len(loader.dataset)


@torch.no_grad()
def evaluate(model, loader, device, criterion):
    model.eval()
    total = 0.0
    preds, trues = [], []
    for xb, yb in loader:
        xb, yb = xb.to(device), yb.to(device)
        p = model(xb)
        total += criterion(p, yb).item() * len(xb)
        preds.append(p.cpu())
        trues.append(yb.cpu())
    return (total / len(loader.dataset),
            torch.cat(preds), torch.cat(trues))


def compute_r2(y_pred_norm, y_true_norm, y_mean, y_std, cols):
    y_pred = (y_pred_norm * y_std + y_mean).numpy()
    y_true = (y_true_norm * y_std + y_mean).numpy()
    out = {}
    for i, col in enumerate(cols):
        err    = y_pred[:, i] - y_true[:, i]
        ss_res = np.sum(err ** 2)
        ss_tot = np.sum((y_true[:, i] - y_true[:, i].mean()) ** 2)
        out[col] = float(1 - ss_res / (ss_tot + 1e-8))
    return out


# ─────────────────────────────────────────────────────────────────────────────
#  Model  (identical to notebook – Sohoite architecture)
# ─────────────────────────────────────────────────────────────────────────────
class Sohoite(nn.Module):
    def __init__(self, input_dim, n_output, dropout):
        super().__init__()
        self.hidden1 = nn.Sequential(
            nn.Linear(input_dim, 64),
            nn.BatchNorm1d(64),
            nn.ReLU(),
            nn.Dropout(dropout),
        )
        self.hidden2 = nn.Sequential(
            nn.Linear(64, 64),
            nn.BatchNorm1d(64),
            nn.ReLU(),
            nn.Dropout(dropout),
        )
        self.skip  = nn.Linear(input_dim, 64)
        self.heads = nn.ModuleList([nn.Linear(64, 1) for _ in range(n_output)])
        self._init_weights()

    def _init_weights(self):
        for m in self.modules():
            if isinstance(m, nn.Linear):
                nn.init.kaiming_uniform_(m.weight, nonlinearity='relu')
                if m.bias is not None:
                    nn.init.zeros_(m.bias)

    def forward(self, x):
        h1 = self.hidden1(x)
        h2 = self.hidden2(h1) + self.skip(x)
        return torch.cat([head(h2) for head in self.heads], dim=1)


# ─────────────────────────────────────────────────────────────────────────────
#  Data loading & preprocessing  (identical to notebook)
# ─────────────────────────────────────────────────────────────────────────────
def load_data(file_name):
    notebook_dir = os.path.dirname(os.path.abspath(__file__))
    file_path    = os.path.join(notebook_dir, 'Data', f'{file_name}.xlsx')

    wb_data = pxl.load_workbook(file_path, data_only=True)
    ws_data = wb_data.active
    headers = [ws_data.cell(1, col).value
               for col in range(1, ws_data.max_column + 1)]
    data    = list(ws_data.iter_rows(min_row=2, values_only=True))
    df      = pd.DataFrame(data, columns=headers)
    df      = df.drop(columns=[None], errors='ignore').dropna()

    # IQR outlier removal (cap at 5 %)
    outlier_cols  = df.columns[:5]
    Q1, Q3        = df[outlier_cols].quantile(0.25), df[outlier_cols].quantile(0.75)
    IQR           = Q3 - Q1
    low_dist      = (Q1 - 1.5 * IQR - df[outlier_cols]).clip(lower=0)
    high_dist     = (df[outlier_cols] - Q3 - 1.5 * IQR).clip(lower=0)
    scores        = (low_dist + high_dist).sum(axis=1)
    max_remove    = int(len(df) * 0.05)
    potential     = scores[scores > 0].sort_values(ascending=False)
    if len(potential) > 0:
        drop_idx = potential.head(min(len(potential), max_remove)).index
        df = df.drop(drop_idx).reset_index(drop=True)

    input_col  = ['r', 'e', 'l', 'Ls', 'Height', 'Width',
                  'Density', 'Pin dia', 'RPM', 'Fbox']
    force_cols = ['|P1| Max', '|B0| Max', 'FOS', 'Torque', 'QRR', 'Power', 'dx']

    x = torch.tensor(df[input_col].values,  dtype=torch.float32)
    y = torch.tensor(df[force_cols].values, dtype=torch.float32)
    return x, y, input_col, force_cols


# ─────────────────────────────────────────────────────────────────────────────
#  Single training run
# ─────────────────────────────────────────────────────────────────────────────
def run_experiment(x, y, y_mean, y_std, x_mean, x_std,
                   train_idx, val_idx, test_idx,
                   force_cols, device,
                   loss_name, criterion, dropout, epochs, opt_name):

    torch.manual_seed(SEED)

    # Normalise
    x_tr, y_tr   = x[train_idx], y[train_idx]
    x_val, y_val = x[val_idx],   y[val_idx]
    x_te,  y_te  = x[test_idx],  y[test_idx]

    x_tr_n  = (x_tr  - x_mean) / x_std
    x_val_n = (x_val - x_mean) / x_std
    x_te_n  = (x_te  - x_mean) / x_std
    y_tr_n  = (y_tr  - y_mean) / y_std
    y_val_n = (y_val - y_mean) / y_std
    y_te_n  = (y_te  - y_mean) / y_std

    val_ds  = Validation(x_val_n, y_val_n)
    test_ds = Testing(x_te_n,  y_te_n)
    val_loader  = DataLoader(val_ds,  batch_size=BATCH_SIZE, shuffle=False)
    test_loader = DataLoader(test_ds, batch_size=BATCH_SIZE, shuffle=False)

    model = Sohoite(x.shape[1], len(force_cols), dropout).to(device)

    if opt_name == "AdamW":
        optimizer = torch.optim.AdamW(model.parameters(), lr=1e-4, weight_decay=5e-5)
    else:  # Adam
        optimizer = torch.optim.Adam(model.parameters(), lr=1e-4)

    best_val, best_state = float('inf'), None
    no_improve = 0

    for epoch in range(1, epochs + 1):
        x_mix, y_mix = mixup(x_tr_n, y_tr_n, alpha=ALPHA, seed=epoch)
        train_ds = Training(x_mix, y_mix)
        train_loader, _, _ = get_dataloaders(
            train_ds, val_ds, test_ds,
            batch_size=BATCH_SIZE, num_workers=0, seed=epoch)

        train_epoch(model, train_loader, optimizer, device, criterion)
        val_loss, _, _ = evaluate(model, val_loader, device, criterion)

        if val_loss < best_val:
            best_val   = val_loss
            best_state = {k: v.clone() for k, v in model.state_dict().items()}
            no_improve = 0
        else:
            no_improve += 1

        if no_improve >= PATIENCE:
            break

    model.load_state_dict(best_state)
    _, test_pred, test_true = evaluate(model, test_loader, device, criterion)
    r2_dict = compute_r2(test_pred, test_true, y_mean, y_std, force_cols)
    mean_r2 = float(np.mean(list(r2_dict.values())))
    return r2_dict, mean_r2


# ─────────────────────────────────────────────────────────────────────────────
#  Excel report builder
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(results, force_cols, out_path):
    wb = Workbook()

    # ── colour palette ──────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
    SUB_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
    ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue
    BEST_FILL  = PatternFill("solid", fgColor="E2EFDA")   # soft green
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    SUB_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DATA_FONT  = Font(name="Arial", size=10)
    BOLD_FONT  = Font(name="Arial", bold=True, size=10)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── helpers ─────────────────────────────────────────────────────────────
    def hdr(ws, row, col, val, fill=HDR_FILL, font=HDR_FONT):
        c = ws.cell(row, col, val)
        c.fill, c.font, c.alignment, c.border = fill, font, CENTER, BORDER
        return c

    def dat(ws, row, col, val, fill=WHITE_FILL, bold=False, fmt=None):
        c = ws.cell(row, col, val)
        c.fill   = fill
        c.font   = BOLD_FONT if bold else DATA_FONT
        c.alignment = CENTER
        c.border = BORDER
        if fmt:
            c.number_format = fmt
        return c

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 1 – Full Results
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Full Results"
    ws.freeze_panes = "A3"

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=5 + len(force_cols))
    title_cell = ws.cell(1, 1, "Jodoigne – Hyperparameter Grid Search Results")
    title_cell.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = CENTER

    # Column headers (row 2)
    cols_hdr = ["Loss Function", "Dropout", "Max Epochs", "Optimizer",
                "Mean R²"] + force_cols
    for ci, h in enumerate(cols_hdr, 1):
        hdr(ws, 2, ci, h)

    # Data rows
    for ri, row in enumerate(results, 3):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        dat(ws, ri, 1, row["loss"],      fill, bold=True)
        dat(ws, ri, 2, row["dropout"],   fill, fmt="0.00")
        dat(ws, ri, 3, row["epochs"],    fill)
        dat(ws, ri, 4, row["optimizer"], fill)
        dat(ws, ri, 5, row["mean_r2"],   fill, fmt="0.0000")
        for ci, col in enumerate(force_cols, 6):
            dat(ws, ri, ci, row["r2"][col], fill, fmt="0.0000")

    # Highlight best mean R² row
    best_mean = max(results, key=lambda r: r["mean_r2"])
    best_ri   = results.index(best_mean) + 3
    for ci in range(1, 6 + len(force_cols)):
        ws.cell(best_ri, ci).fill = BEST_FILL
        ws.cell(best_ri, ci).font = BOLD_FONT

    # Conditional colour-scale on Mean R² column (col 5)
    last_data_row = 2 + len(results)
    r2_range = f"E3:E{last_data_row}"
    ws.conditional_formatting.add(
        r2_range,
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 12
    for ci in range(6, 6 + len(force_cols)):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 2 – Summary by Factor
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary by Factor")

    def section_title(ws, row, col, title, span):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
        c = ws.cell(row, col, title)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = CENTER
        c.border    = BORDER

    # ── Loss Function summary ────────────────────────────────────────────────
    section_title(ws2, 1, 1, "Mean R² by Loss Function", 2)
    hdr(ws2, 2, 1, "Loss Function"); hdr(ws2, 2, 2, "Avg Mean R²")
    r = 3
    for loss_name in LOSS_FNS:
        vals = [x["mean_r2"] for x in results if x["loss"] == loss_name]
        dat(ws2, r, 1, loss_name, bold=True)
        dat(ws2, r, 2, np.mean(vals), fmt="0.0000")
        r += 1

    # ── Dropout summary ──────────────────────────────────────────────────────
    section_title(ws2, 1, 4, "Mean R² by Dropout", 2)
    hdr(ws2, 2, 4, "Dropout"); hdr(ws2, 2, 5, "Avg Mean R²")
    r = 3
    for d in DROPOUTS:
        vals = [x["mean_r2"] for x in results if x["dropout"] == d]
        dat(ws2, r, 4, d, fmt="0.00")
        dat(ws2, r, 5, np.mean(vals), fmt="0.0000")
        r += 1

    # ── Epoch summary ────────────────────────────────────────────────────────
    section_title(ws2, 1, 7, "Mean R² by Max Epochs", 2)
    hdr(ws2, 2, 7, "Max Epochs"); hdr(ws2, 2, 8, "Avg Mean R²")
    r = 3
    for ep in EPOCH_LIST:
        vals = [x["mean_r2"] for x in results if x["epochs"] == ep]
        dat(ws2, r, 7, ep)
        dat(ws2, r, 8, np.mean(vals), fmt="0.0000")
        r += 1

    # ── Optimizer summary ────────────────────────────────────────────────────
    section_title(ws2, 1, 10, "Mean R² by Optimizer", 2)
    hdr(ws2, 2, 10, "Optimizer"); hdr(ws2, 2, 11, "Avg Mean R²")
    r = 3
    for opt in OPTIMIZERS:
        vals = [x["mean_r2"] for x in results if x["optimizer"] == opt]
        dat(ws2, r, 10, opt, bold=True)
        dat(ws2, r, 11, np.mean(vals), fmt="0.0000")
        r += 1

    # ── Best configuration box ───────────────────────────────────────────────
    best = max(results, key=lambda x: x["mean_r2"])
    section_title(ws2, 8, 1, "🏆  Best Configuration", 6)
    labels  = ["Loss", "Dropout", "Max Epochs", "Optimizer", "Mean R²"]
    vals_b  = [best["loss"], best["dropout"], best["epochs"],
               best["optimizer"], best["mean_r2"]]
    fmts    = [None, "0.00", None, None, "0.0000"]
    for ci, (lbl, val, fmt) in enumerate(zip(labels, vals_b, fmts), 1):
        hdr(ws2, 9, ci, lbl, fill=SUB_FILL, font=SUB_FONT)
        c = ws2.cell(10, ci, val)
        c.fill      = BEST_FILL
        c.font      = BOLD_FONT
        c.alignment = CENTER
        c.border    = BORDER
        if fmt:
            c.number_format = fmt

    for ci in range(1, 12):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    # ════════════════════════════════════════════════════════════════════════
    #  Sheet 3 – Per-Target R²
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Per-Target R²")

    # Header
    ws3.merge_cells(start_row=1, start_column=1,
                    end_row=1, end_column=4 + len(force_cols))
    tc = ws3.cell(1, 1, "Per-Target R² for Every Configuration")
    tc.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = CENTER

    hdrs3 = ["Loss Function", "Dropout", "Epochs", "Optimizer"] + force_cols
    for ci, h in enumerate(hdrs3, 1):
        hdr(ws3, 2, ci, h)

    for ri, row in enumerate(results, 3):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        dat(ws3, ri, 1, row["loss"],      fill, bold=True)
        dat(ws3, ri, 2, row["dropout"],   fill, fmt="0.00")
        dat(ws3, ri, 3, row["epochs"],    fill)
        dat(ws3, ri, 4, row["optimizer"], fill)
        for ci, col in enumerate(force_cols, 5):
            dat(ws3, ri, ci, row["r2"][col], fill, fmt="0.0000")

    last_row3 = 2 + len(results)
    for ci in range(5, 5 + len(force_cols)):
        col_letter = get_column_letter(ci)
        ws3.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{last_row3}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 11
    for ci in range(5, 5 + len(force_cols)):
        ws3.column_dimensions[get_column_letter(ci)].width = 13

    ws3.row_dimensions[1].height = 22
    ws3.row_dimensions[2].height = 36
    ws3.freeze_panes = "A3"

    wb.save(out_path)
    print(f"\n✅  Results saved → {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # Device
    device = torch.device(
        'mps'  if torch.backends.mps.is_available() else
        'cuda' if torch.cuda.is_available()         else
        'cpu'
    )
    print(f"Device : {device}")

    # Load & prep data once
    x, y, input_col, force_cols = load_data(FILE_NAME)

    train_idx, val_idx, test_idx = split_data(x, y, SEED)

    x_tr = x[train_idx]; y_tr = y[train_idx]
    x_mean = x_tr.mean(dim=0)
    x_std  = x_tr.std(dim=0).clamp(min=1e-8)
    y_mean = y_tr.mean(dim=0)
    y_std  = y_tr.std(dim=0).clamp(min=1e-8)

    # Build grid
    grid = list(itertools.product(
        LOSS_FNS.items(),   # (name, criterion)
        DROPOUTS,
        EPOCH_LIST,
        OPTIMIZERS,
    ))
    total = len(grid)
    print(f"\nTotal experiments : {total}\n")
    print(f"{'#':>4}  {'Loss':<12} {'Dropout':>8} {'Epochs':>8} "
          f"{'Optimizer':<8}  Mean R²")
    print("─" * 62)

    results = []
    for idx, ((loss_name, criterion), dropout, epochs, opt_name) in enumerate(grid, 1):
        t0 = time.time()
        r2_dict, mean_r2 = run_experiment(
            x, y, y_mean, y_std, x_mean, x_std,
            train_idx, val_idx, test_idx,
            force_cols, device,
            loss_name, criterion, dropout, epochs, opt_name,
        )
        elapsed = time.time() - t0
        print(f"{idx:>4}  {loss_name:<12} {dropout:>8.2f} {epochs:>8} "
              f"{opt_name:<8}  {mean_r2:.4f}   ({elapsed:.0f}s)")

        results.append({
            "loss"      : loss_name,
            "dropout"   : dropout,
            "epochs"    : epochs,
            "optimizer" : opt_name,
            "mean_r2"   : mean_r2,
            "r2"        : r2_dict,
        })

    # Save Excel
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "hyperparameter_results.xlsx",
    )
    build_excel(results, force_cols, out_path)

    # Print podium
    ranked = sorted(results, key=lambda r: r["mean_r2"], reverse=True)
    print("\n─── Top 5 configurations ───────────────────────────────────")
    for i, r in enumerate(ranked[:5], 1):
        print(f"  #{i}  {r['loss']:<12}  dropout={r['dropout']:.2f}  "
              f"epochs={r['epochs']}  {r['optimizer']:<6}  "
              f"Mean R²={r['mean_r2']:.4f}")
    print("─" * 62)


if __name__ == "__main__":
    main()
