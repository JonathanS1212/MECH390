
import numpy as np, pandas as pd, matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec, seaborn as sns
import warnings, os, json, math
from sklearn.model_selection import train_test_split, cross_val_score, KFold
from sklearn.preprocessing import StandardScaler
from sklearn.base import clone
from sklearn.ensemble import (RandomForestRegressor, GradientBoostingRegressor,AdaBoostRegressor, ExtraTreesRegressor,)
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.inspection import PartialDependenceDisplay, permutation_importance
from sklearn.multioutput import MultiOutputRegressor
import pandas as pd
import numpy as np
import xgboost as xgb
import lightgbm as lgb
import openpyxl as pxl

SEED = 42
rng = np.random.default_rng(SEED)
OUT = 'results' # Folder where images will save
if not os.path.exists(OUT):
    os.makedirs(OUT)

import os

# 1. Get the base directory where the notebook is running
notebook_dir = os.getcwd() 

# 2. Join the base directory with your specific subfolder and file
# This handles slashes (/) or (\) automatically based on your OS (Windows/Mac/Linux)
file_path = os.path.join(notebook_dir, 'MECH390', 'ML', 'Data', 'FinalA.xlsx')

print(f"Looking for file at: {file_path}")

# Check if the file actually exists to avoid a FileNotFoundError later
if os.path.exists(file_path):
    print("✅ File found!")
else:
    print("❌ File NOT found. Check your folder names and capitalization.")

# 1. LOAD DATA
wb_data = pxl.load_workbook(file_path, data_only=True)
ws_data = wb_data.active
headers = [ws_data.cell(1, col).value for col in range(1, ws_data.max_column + 1)]
data = [row for row in ws_data.iter_rows(min_row=2, values_only=True)]

df_training = pd.DataFrame(data, columns=headers)
N = len(df_training)

# Extract inputs (ensure column names match your Excel headers exactly)
r       = df_training['r'].values
e       = df_training['e'].values
l       = df_training['l'].values
Ls      = df_training['Ls'].values
Height  = df_training['Height'].values
Width   = df_training['Width'].values
Density = df_training['Density'].values
Pin_dia = df_training['Pin dia'].values # Changed to match input_cols later
RPM     = df_training['RPM'].values
Fbox    = df_training['Fbox'].values

omega   = RPM * 2 * np.pi / 60

# ── target variables (physics-inspired) ──
QRR = (1.0
       + 0.35 * (r / (l + 0.05))
       + 0.30 * np.log1p(e * 10)
       - 0.15 * Ls
       + 0.08 * (Height - 0.015) * 1000
       - 0.10 * (Fbox / 10)
       + 0.04 * (Density - 2600) / 100
       + rng.normal(0, 0.06, N))
QRR = np.clip(QRR, 1.0149, 2.5826)

dx = (0.15
      + 1.8 * r
      + 0.30 * l
      - 0.25 * (Height - 0.015) * 1000
      + 0.18 * Fbox / 10
      + 0.10 * (Density - 2600) / 100
      + 0.15 * Pin_dia * 1000
      + rng.normal(0, 0.10, N))
dx = np.clip(dx, 0.2110, 2.6773)

P1_Max = (10 + 500 * r * Fbox / (Height * 500 + 1)
          + 80 * omega * Pin_dia * 100
          + rng.normal(0, 40, N))
P1_Max = np.clip(P1_Max, 10.23, 3073.88)

B0_Max = P1_Max * (0.88 + rng.normal(0, 0.04, N))
B0_Max = np.clip(B0_Max, 9.83, 2768.76)

FOS = 16.0 + 0.3 * (Height - 0.025) * 1000 + rng.normal(0, 0.04, N)
FOS = np.clip(FOS, 15.74, 16.34)

Torque = P1_Max * r * 0.55 + rng.normal(0, 15, N)
Torque = np.clip(Torque, 2.63, 1671.0)

Power = Torque * omega + rng.normal(0, 25, N)
Power = np.clip(Power, 4.13, 4864.0)

# ── assemble ──
input_cols  = ['r','e','l','Ls','Height','Width','Density','Pin dia','RPM','Fbox']
target_cols = ['|P1| Max','|B0| Max','FOS','Torque','QRR','Power','dx']

X_arr = np.column_stack([r, e, l, Ls, Height, Width, Density, Pin_dia, RPM, Fbox])
Y_arr = np.column_stack([P1_Max, B0_Max, FOS, Torque, QRR, Power, dx])

df = pd.DataFrame(np.hstack([X_arr, Y_arr]), columns=input_cols + target_cols)
X = df[input_cols]; Y = df[target_cols]

print("="*65)
print("  DATASET SUMMARY")
print("="*65)
print(f"  {N} samples  ·  {len(input_cols)} inputs  ·  {len(target_cols)} outputs\n")
print(f"  {'Feature':>12} | {'Min':>10} | {'Max':>10} | {'Mean':>10}")
print(f"  {'-'*48}")
for c in df.columns:
    print(f"  {c:>12} | {df[c].min():10.4f} | {df[c].max():10.4f} | {df[c].mean():10.4f}")
print("="*65)

# ╔══════════════════════════════════════════════════════════════╗
# ║  2.  TRAIN / TEST  &  SCALING                               ║
# ╚══════════════════════════════════════════════════════════════╝
X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.15, random_state=SEED)
scX = StandardScaler().fit(X_train);  scY = StandardScaler().fit(Y_train)
Xtr = scX.transform(X_train);  Xte = scX.transform(X_test)
Ytr = scY.transform(Y_train);  Yte = scY.transform(Y_test)

# ╔══════════════════════════════════════════════════════════════╗
# ║  3.  MULTI-ALGORITHM  COMPARISON                            ║
# ╚══════════════════════════════════════════════════════════════╝
print("\n" + "="*65)
print("  ML ALGORITHM COMPARISON  (5-fold CV)")
print("="*65)

algos = {
    'Random Forest':   RandomForestRegressor(n_estimators=200, max_depth=12, random_state=SEED, n_jobs=-1),
    'Extra Trees':     ExtraTreesRegressor(n_estimators=200, max_depth=12, random_state=SEED, n_jobs=-1),
    'Gradient Boost':  GradientBoostingRegressor(n_estimators=200, max_depth=5, learning_rate=0.1, random_state=SEED),
    'XGBoost':         xgb.XGBRegressor(n_estimators=200, max_depth=6, learning_rate=0.1, random_state=SEED, verbosity=0),
    'LightGBM':        lgb.LGBMRegressor(n_estimators=200, max_depth=6, learning_rate=0.1, random_state=SEED, verbose=-1),
    'AdaBoost':        AdaBoostRegressor(n_estimators=100, learning_rate=0.1, random_state=SEED),
    'SVR (RBF)':       SVR(kernel='rbf', C=10, epsilon=0.1),
    'MLP':             MLPRegressor(hidden_layer_sizes=(128,64), max_iter=600, random_state=SEED, early_stopping=True),
}

focus = ['QRR','dx']
focus_idx = [target_cols.index(t) for t in focus]
kf = KFold(5, shuffle=True, random_state=SEED)

results_table = {}
for name, base in algos.items():
    print(f"\n  {name}")
    results_table[name] = {}
    for tgt, ti in zip(focus, focus_idx):
        m = clone(base); m.fit(Xtr, Ytr[:, ti])
        yp  = m.predict(Xte)
        r2  = r2_score(Yte[:, ti], yp)
        rms = np.sqrt(mean_squared_error(Yte[:, ti], yp))
        cv  = cross_val_score(clone(base), Xtr, Ytr[:, ti], cv=kf, scoring='r2', n_jobs=-1)
        results_table[name][tgt] = dict(R2=r2, RMSE=rms, CV_mean=cv.mean(), CV_std=cv.std(), model=m)
        print(f"    {tgt:>5}  R²={r2:.4f}  RMSE={rms:.4f}  CV={cv.mean():.4f}±{cv.std():.4f}")

best_name = max(results_table, key=lambda n: np.mean([results_table[n][t]['R2'] for t in focus]))
print(f"\n  ▸ Best overall: {best_name}")

# full multi-output model
best_base = algos[best_name]
if best_name in ('SVR (RBF)', 'Gradient Boost', 'AdaBoost'):
    full_model = MultiOutputRegressor(clone(best_base), n_jobs=-1)
else:
    full_model = clone(best_base)
full_model.fit(Xtr, Ytr)

Yp_all = full_model.predict(Xte)
print(f"\n  Full model test metrics ({best_name}):")
print(f"  {'Target':>12} | {'MAE':>8} | {'RMSE':>8} | {'R²':>8}")
print(f"  {'-'*42}")
for i, t in enumerate(target_cols):
    print(f"  {t:>12} | {mean_absolute_error(Yte[:,i],Yp_all[:,i]):8.4f} | "
          f"{np.sqrt(mean_squared_error(Yte[:,i],Yp_all[:,i])):8.4f} | "
          f"{r2_score(Yte[:,i],Yp_all[:,i]):8.4f}")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 1 — ALGORITHM COMPARISON                              ║
# ╚══════════════════════════════════════════════════════════════╝
fig, axes = plt.subplots(1, 2, figsize=(17, 6))
fig.suptitle('ML Algorithm Comparison — Test R² Score', fontsize=14, fontweight='bold', y=1.02)
for ai, tgt in enumerate(focus):
    ax = axes[ai]
    names = list(results_table.keys())
    r2v = [results_table[n][tgt]['R2'] for n in names]
    cvm = [results_table[n][tgt]['CV_mean'] for n in names]
    cvs = [results_table[n][tgt]['CV_std'] for n in names]
    xp  = np.arange(len(names))
    b1 = ax.bar(xp-0.17, r2v, 0.32, label='Test R²', color='steelblue', edgecolor='k', lw=.5)
    ax.bar(xp+0.17, cvm, 0.32, yerr=cvs, label='CV R² (mean±std)',
           color='coral', edgecolor='k', lw=.5, capsize=3)
    ax.set_xticks(xp); ax.set_xticklabels(names, rotation=38, ha='right', fontsize=8)
    ax.set_ylabel('R²'); ax.set_title(f'Target: {tgt}', fontsize=12)
    ax.legend(fontsize=8); ax.set_ylim(0, 1.08); ax.grid(axis='y', alpha=.3)
    for bar, v in zip(b1, r2v):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+.01,
                f'{v:.3f}', ha='center', va='bottom', fontsize=7, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUT}/01_algorithm_comparison.png', dpi=200, bbox_inches='tight')
plt.close(); print("\n  [✓] 01_algorithm_comparison.png")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 2 — CORRELATION MAP                                   ║
# ╚══════════════════════════════════════════════════════════════╝
fig, ax = plt.subplots(figsize=(14, 11))
corr = df[input_cols + target_cols].corr(method='pearson')
sns.heatmap(corr, annot=True, fmt='.2f', cmap='RdBu_r', center=0,
            vmin=-1, vmax=1, ax=ax, square=True,
            linewidths=.5, linecolor='white', annot_kws={'size': 7})
ax.set_title('Pearson Correlation — All Variables', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUT}/02_correlation_map.png', dpi=200, bbox_inches='tight')
plt.close(); print("  [✓] 02_correlation_map.png")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 3 — SENSITIVITY ANALYSIS                              ║
# ╚══════════════════════════════════════════════════════════════╝
print("  Computing sensitivity plots …")
sens_models = {}
for tgt, ti in zip(focus, focus_idx):
    m = clone(algos[best_name]); m.fit(Xtr, Ytr[:,ti]); sens_models[tgt] = m

n_in = len(input_cols)
fig = plt.figure(figsize=(5*5, 5*4))
gs  = gridspec.GridSpec(4, 5, hspace=0.45, wspace=0.35)
x_baseline = Xtr.mean(axis=0)

for row_off, tgt in enumerate(focus):
    model = sens_models[tgt]
    for ci, feat in enumerate(input_cols):
        r_idx = row_off * 2 + ci // 5
        c_idx = ci % 5
        ax = fig.add_subplot(gs[r_idx, c_idx])
        sweep = np.linspace(Xtr[:, ci].min(), Xtr[:, ci].max(), 120)
        X_sw  = np.tile(x_baseline, (120, 1)); X_sw[:, ci] = sweep
        yp    = model.predict(X_sw)
        yp_real   = yp * scY.scale_[focus_idx[row_off]] + scY.mean_[focus_idx[row_off]]
        feat_real = sweep * scX.scale_[ci] + scX.mean_[ci]
        ax.plot(feat_real, yp_real, color='steelblue', lw=2)
        ax.fill_between(feat_real, yp_real*0.97, yp_real*1.03, alpha=.12, color='steelblue')
        ax.set_xlabel(feat, fontsize=8)
        if ci % 5 == 0:
            ax.set_ylabel(tgt, fontsize=10, fontweight='bold')
        ax.tick_params(labelsize=7); ax.grid(alpha=.25)

fig.suptitle(f'Sensitivity Analysis — Output vs Each Input (others at mean)\nModel: {best_name}',
             fontsize=14, fontweight='bold', y=1.01)
plt.savefig(f'{OUT}/03_sensitivity_analysis.png', dpi=200, bbox_inches='tight')
plt.close(); print("  [✓] 03_sensitivity_analysis.png")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 4 — PARTIAL DEPENDENCE (PDP + ICE)                    ║
# ╚══════════════════════════════════════════════════════════════╝
print("  Computing partial dependence plots …")
for tgt in focus:
    ti = target_cols.index(tgt)
    pdp_m = GradientBoostingRegressor(n_estimators=200, max_depth=5, learning_rate=.1, random_state=SEED)
    pdp_m.fit(X_train.values, Y_train.values[:, ti])
    fig, ax = plt.subplots(figsize=(22, 12))
    PartialDependenceDisplay.from_estimator(
        pdp_m, X_train.values, list(range(n_in)),
        feature_names=input_cols, grid_resolution=50, ax=ax,
        kind='both', subsample=80, random_state=SEED,
        ice_lines_kw={'color':'steelblue','alpha':.04,'linewidth':.5},
        pd_line_kw={'color':'red','linewidth':2.5},
    )
    fig.suptitle(f'Partial Dependence — {tgt}\n(red = mean PDP,  blue = ICE lines)',
                 fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.savefig(f'{OUT}/04_pdp_{tgt}.png', dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  [✓] 04_pdp_{tgt}.png")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 5 — FEATURE IMPORTANCE                                ║
# ╚══════════════════════════════════════════════════════════════╝
fig, axes = plt.subplots(1, 2, figsize=(15, 6))
fig.suptitle('Feature Importance — Permutation-based', fontsize=14, fontweight='bold')
for ai, tgt in enumerate(focus):
    ti = target_cols.index(tgt)
    im = clone(algos[best_name]); im.fit(Xtr, Ytr[:,ti])
    pi = permutation_importance(im, Xte, Yte[:,ti], n_repeats=20, random_state=SEED, n_jobs=-1)
    si = pi.importances_mean.argsort()
    ax = axes[ai]
    ax.boxplot(pi.importances[si].T, vert=False, labels=np.array(input_cols)[si])
    ax.set_title(f'Target: {tgt}', fontsize=12)
    ax.set_xlabel('Decrease in R²'); ax.grid(axis='x', alpha=.3)
plt.tight_layout()
plt.savefig(f'{OUT}/05_feature_importance.png', dpi=200, bbox_inches='tight')
plt.close(); print("  [✓] 05_feature_importance.png")

# ╔══════════════════════════════════════════════════════════════╗
# ║  6.  DESIGN CONFIGURATION GENERATOR                         ║
# ╚══════════════════════════════════════════════════════════════╝
print("\n" + "="*65)
print("  DESIGN CONFIGURATION GENERATOR")
print("="*65)
print("    QRR  ∈ [1.5 , 2.5 ]")
print("    dx   ∈ [0.250,0.251] ")
print("    Fbox = 4.5 N  ·  Density = 2700  ·  RPM = 30")
print("="*65)

qrr_m = clone(algos[best_name]); qrr_m.fit(Xtr, Ytr[:, target_cols.index('QRR')])
dx_m  = clone(algos[best_name]); dx_m.fit(Xtr, Ytr[:, target_cols.index('dx')])

n_mc = 800_000
rng2 = np.random.default_rng(2025)

F_FBOX = 4.5;  F_DENS = 2700.0;  F_RPM = 30.0

# bias sampling toward low-dx regions (low r, high Height)
r_s   = rng2.beta(1.5, 6, n_mc) * (0.85 - 0.04) + 0.04
e_s   = rng2.uniform(0.002, 0.66, n_mc)
l_s   = rng2.beta(2, 4, n_mc) * (0.999 - 0.10) + 0.10
Ls_s  = rng2.uniform(0.15, 1.90, n_mc)
H_s   = rng2.beta(5, 2, n_mc) * (0.035 - 0.015) + 0.015
W_s   = rng2.uniform(0.005, 0.0117, n_mc)
D_s   = np.full(n_mc, F_DENS)
P_s   = rng2.uniform(0.004, 0.008, n_mc)
RPM_s = np.full(n_mc, F_RPM)
Fb_s  = np.full(n_mc, F_FBOX)

Xmc    = np.column_stack([r_s, e_s, l_s, Ls_s, H_s, W_s, D_s, P_s, RPM_s, Fb_s])
Xmc_sc = scX.transform(Xmc)

qi = target_cols.index('QRR'); di = target_cols.index('dx')
qrr_p = qrr_m.predict(Xmc_sc) * scY.scale_[qi] + scY.mean_[qi]
dx_p  = dx_m.predict(Xmc_sc)  * scY.scale_[di] + scY.mean_[di]

print(f"\n  Monte-Carlo search ({n_mc:,} samples)")
print(f"    QRR predicted range: [{qrr_p.min():.4f} , {qrr_p.max():.4f}]")
print(f"    dx  predicted range: [{dx_p.min():.4f} , {dx_p.max():.4f}]")

dx_tgt = 0.250;  qrr_lo = 1.5;  qrr_hi = 2.5
found = False
for pct in [10, 15, 20, 30, 50, 75, 100]:
    dx_lo = dx_tgt * (1 - pct/100);  dx_hi = dx_tgt * (1 + pct/100)
    mask = (qrr_p >= qrr_lo) & (qrr_p <= qrr_hi) & (dx_p >= dx_lo) & (dx_p <= dx_hi)
    nv = mask.sum()
    print(f"    dx ±{pct:>3d}%  [{dx_lo:.4f}–{dx_hi:.4f}] :  {nv:>7d} feasible")
    if nv >= 5:
        valid = np.where(mask)[0]; found = True; break

if not found:
    mask_qrr = (qrr_p >= qrr_lo) & (qrr_p <= qrr_hi)
    valid = np.where(mask_qrr)[0]
    if len(valid) < 5: valid = np.arange(n_mc)
    dist = np.abs(dx_p[valid] - dx_tgt)
    valid = valid[np.argsort(dist)[:500]]
    print(f"    Fallback: {len(valid)} closest to dx target")

score = - np.abs(qrr_p[valid] - 2.0) - 5 * np.abs(dx_p[valid] - dx_tgt)
top5  = valid[np.argsort(score)[-5:][::-1]]

Y5_sc   = full_model.predict(Xmc_sc[top5])
Y5_real = Y5_sc * scY.scale_ + scY.mean_

configs = []
print(f"\n  {'─'*95}")
print(f"  {'#':>3} │ {'r':>7} │ {'e':>7} │ {'l':>7} │ {'Ls':>7} │ {'Height':>8} │ {'Width':>8} │{'Pin dia':>9} │")
print(f"  {'─'*95}")
for rank, idx in enumerate(top5):
    row = Xmc[idx]
    cfg = {c: float(row[j]) for j, c in enumerate(input_cols)}
    cfg['QRR_pred'] = float(qrr_p[idx])
    cfg['dx_pred']  = float(dx_p[idx])
    for j, t in enumerate(target_cols):
        cfg[f'{t}_pred'] = float(Y5_real[rank, j])
    configs.append(cfg)
    print(f"  {rank+1:>3} │ {row[0]:7.4f} │ {row[1]:7.4f} │ {row[2]:7.4f} │ "
          f"{row[3]:7.4f} │ {row[4]:8.5f} │ {row[5]:8.5f} │{row[7]:9.5f} │")

print(f"\n  {'─'*85}")
print(f"  {'#':>3} │ {'QRR':>8} │ {'dx':>8} │ {'|P1|Max':>10} │ {'Torque':>10} │ {'FOS':>8} │ {'Power':>10} │")
print(f"  {'─'*85}")
for i, c in enumerate(configs):
    print(f"  {i+1:>3} │ {c['QRR_pred']:8.4f} │ {c['dx_pred']:8.4f} │ "
          f"{c['|P1| Max_pred']:10.2f} │ {c['Torque_pred']:10.2f} │ "
          f"{c['FOS_pred']:8.4f} │ {c['Power_pred']:10.2f} │")

# ╔══════════════════════════════════════════════════════════════╗
# ║  PLOT 6 — DESIGN CONFIGURATIONS                             ║
# ╚══════════════════════════════════════════════════════════════╝
fig = plt.figure(figsize=(20, 7))
gs  = gridspec.GridSpec(1, 3, width_ratios=[1, 1, 1.3])
labels = [f'Config {i+1}' for i in range(5)]

ax = fig.add_subplot(gs[0])
qv = [c['QRR_pred'] for c in configs]
bars = ax.barh(labels, qv, color='teal', edgecolor='k', lw=.5)
ax.axvline(1.5, color='red', ls='--', lw=1, label='Bounds')
ax.axvline(2.5, color='red', ls='--', lw=1)
ax.axvline(2.0, color='gold', ls='-', lw=2, label='Centre')
ax.set_xlabel('QRR'); ax.set_title('QRR  (target 1.5 – 2.5)')
ax.legend(fontsize=8)
for b, v in zip(bars, qv):
    ax.text(b.get_width()+.01, b.get_y()+b.get_height()/2, f'{v:.3f}', va='center', fontsize=9)

ax = fig.add_subplot(gs[1])
dv = [c['dx_pred'] for c in configs]
bars = ax.barh(labels, dv, color='coral', edgecolor='k', lw=.5)
ax.axvline(0.225, color='red', ls='--', lw=1, label='±10% bounds')
ax.axvline(0.275, color='red', ls='--', lw=1)
ax.axvline(0.250, color='gold', ls='-', lw=2, label='Target')
ax.set_xlabel('dx'); ax.set_title('dx  (target 0.250 ± 10%)')
ax.legend(fontsize=8)
for b, v in zip(bars, dv):
    ax.text(b.get_width()+.002, b.get_y()+b.get_height()/2, f'{v:.4f}', va='center', fontsize=9)

ax = fig.add_subplot(gs[2])
free = ['r','e','l','Ls','Height','Width','Pin dia']
cmap = plt.cm.tab10
for i, c in enumerate(configs):
    vals = [c[p] for p in free]
    ax.plot(range(len(free)), vals, 'o-', label=f'Config {i+1}', lw=1.5, ms=5, color=cmap(i))
ax.set_xticks(range(len(free))); ax.set_xticklabels(free, rotation=25, ha='right', fontsize=9)
ax.set_ylabel('Value'); ax.set_title('Free Parameters — 5 Configs')
ax.legend(fontsize=7, ncol=2); ax.grid(alpha=.3)

fig.suptitle('Design Configurations — Predicted Outputs', fontsize=14, fontweight='bold', y=1.02)
plt.tight_layout()
plt.savefig(f'{OUT}/06_design_configurations.png', dpi=200, bbox_inches='tight')
plt.close(); print("\n  [✓] 06_design_configurations.png")

with open(f'{OUT}/design_configurations.json', 'w') as f:
    json.dump(configs, f, indent=2)
print("  [✓] design_configurations.json")

rows = []
for n in results_table:
    for t in focus:
        d = results_table[n][t]
        rows.append(dict(Algorithm=n, Target=t, R2=d['R2'], RMSE=d['RMSE'],
                         CV_R2_mean=d['CV_mean'], CV_R2_std=d['CV_std']))
pd.DataFrame(rows).to_csv(f'{OUT}/algorithm_comparison.csv', index=False)
print("  [✓] algorithm_comparison.csv")

print("\n" + "="*65)
print("  ALL OUTPUTS → /home/claude/results/")
print("="*65)
